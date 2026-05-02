import math
import os
import re
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent.parent


def resolve_dataset_path():
    override = os.environ.get("RIGSYNC_DATASET_PATH")
    if override:
        return Path(override).expanduser()

    search_roots = [
        PROJECT_ROOT,
        Path.home() / "Downloads" / "ise",
        Path.home() / "Downloads",
    ]
    candidate_names = (
        "FinalData.xlsx",
        "ise_data_final_v2.xlsx",
        "iseData.xlsx",
    )

    for search_root in search_roots:
        for candidate_name in candidate_names:
            candidate = search_root / candidate_name
            if candidate.exists():
                return candidate

    return PROJECT_ROOT / "iseData.xlsx"


DATASET_PATH = resolve_dataset_path()

WORKER_ROLE_DEFINITIONS = [
    {"id": "assistant_driller", "label": "Assistant Driller"},
    {"id": "bop_tech", "label": "BOP Tech"},
    {"id": "camp_foreman", "label": "Camp Foreman"},
    {"id": "crane_operator", "label": "Crane Operator"},
    {"id": "derrickman", "label": "Derrickman"},
    {"id": "driller", "label": "Driller"},
    {"id": "electrician", "label": "Electrician"},
    {"id": "floorman", "label": "Floorman"},
    {"id": "forklift_crane_operator", "label": "Forklift/Crane Operator"},
    {"id": "mechanic", "label": "Mechanic"},
    {"id": "operator", "label": "Operator"},
    {"id": "pumpman_mechanic", "label": "Pumpman/Mechanic"},
    {"id": "rigger", "label": "Rigger"},
    {"id": "roustabout", "label": "Roustabout"},
    {"id": "welder", "label": "Welder"},
    {"id": "yard_foreman", "label": "Yard Foreman"},
]

WORKER_ROLE_ALIASES = {
    "asst. driller": "assistant_driller",
    "assistant driller": "assistant_driller",
    "bop tech": "bop_tech",
    "bop techs": "bop_tech",
    "camp foreman": "camp_foreman",
    "crane operator": "crane_operator",
    "crane operators": "crane_operator",
    "derrickman": "derrickman",
    "driller": "driller",
    "electrician": "electrician",
    "electricians": "electrician",
    "floormen": "floorman",
    "floor men": "floorman",
    "forklift/crane operator": "forklift_crane_operator",
    "mechanic": "mechanic",
    "mechanics": "mechanic",
    "operator": "operator",
    "pumpman/mechanic": "pumpman_mechanic",
    "rigger": "rigger",
    "riggers": "rigger",
    "roustabout": "roustabout",
    "roustabouts": "roustabout",
    "welder": "welder",
    "welders": "welder",
    "yard foreman": "yard_foreman",
    "truck driver": "truck_driver",
}

TRUCK_TYPE_ALIASES = {
    "flatbed": "Flat-bed",
    "flat-bed": "Flat-bed",
    "lowbed": "Low-bed",
    "low-bed": "Low-bed",
    "heavyhauler": "Heavy Hauler",
    "heavy hauler": "Heavy Hauler",
}

STARTUP_REUSABLE_IDS = {
    "SU-01",
    "SU-02",
    "SU-03",
    "SU-04",
    "SU-05",
    "SU-06",
    "SU-13",
    "SU-14",
    "SU-15",
}

SHEET_NAME_CANDIDATES = {
    "rig": ("Expanded Loads", "Rig Loads"),
    "startup": ("Expanded Additional Loads", "Additional Needed Loads"),
    "truck": ("Trucks Info",),
}

PHASE_ROW_PATTERN = re.compile(r"^((?:RL|CL|SU)-\d+(?:-L\d+)?)\s*\((RD|RM|RU)\)\s*$", flags=re.IGNORECASE)
WORKBOOK_ROLE_COLUMNS = [
    "driller",
    "assistant_driller",
    "derrickman",
    "floorman",
    "rigger",
    "crane_operator",
    "mechanic",
    "welder",
    "electrician",
    "pumpman_mechanic",
    "bop_tech",
    "forklift_crane_operator",
    "operator",
    "camp_foreman",
    "yard_foreman",
    "roustabout",
    "truck_driver",
]


def normalize_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_int(value):
    text = normalize_text(value)
    if text is None:
        return None
    match = re.search(r"-?\d+", text.replace(",", ""))
    return int(match.group()) if match else None


def normalize_float(value):
    text = normalize_text(value)
    if text is None:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text.replace(",", ""))
    return float(match.group()) if match else None


def parse_duration_minutes(value):
    text = normalize_text(value)
    if not text or text in {"—", "(dist / speed)"}:
        return None
    if "±" in text:
        text = text.split("±", 1)[0].strip()
    match = re.search(r"(\d+(?:\.\d+)?)", text, flags=re.IGNORECASE)
    return int(math.ceil(float(match.group(1)) * 60)) if match else None


def parse_dimension_triplet(value):
    text = normalize_text(value)
    if not text:
        return None
    values = [float(match) for match in re.findall(r"\d+(?:\.\d+)?", text)]
    if len(values) < 3:
        return None
    return {
        "length": values[0],
        "width": values[1],
        "height": values[2],
    }


def parse_weight_tons(value):
    text = normalize_text(value)
    if not text:
        return None
    values = [float(match) for match in re.findall(r"\d+(?:\.\d+)?", text.replace(",", ""))]
    return max(values) if values else None


def parse_yes_no(value):
    text = normalize_text(value)
    return bool(text and text.lower() in {"yes", "true", "1"})


def normalize_truck_type(value):
    text = normalize_text(value)
    if not text:
        return None
    key = re.sub(r"[^a-z]", "", text.lower())
    if key == "fb" or "flatbed" in key:
        key = "flatbed"
    elif key == "lb" or "lowbed" in key or "support" in key:
        key = "lowbed"
    elif key == "hh" or "heavyhaul" in key:
        key = "heavyhauler"
    return TRUCK_TYPE_ALIASES.get(key, text)


def parse_truck_types(value):
    text = normalize_text(value)
    if not text:
        return []
    return [
        normalized
        for normalized in (
            normalize_truck_type(part)
            for part in re.split(r"[/|,;]+", text)
        )
        if normalized
    ]


def parse_dependency_codes(value):
    text = normalize_text(value)
    if not text or text == "—":
        return []
    return [match.upper() for match in re.findall(r"\b(?:RL|CL|SU)-\d+(?:-L\d+)?\b", text, flags=re.IGNORECASE)]


def parse_phase_dependency_refs(value):
    text = normalize_text(value)
    if not text or text == "—":
        return []
    return [
        f"{code.upper()} ({phase.upper()})"
        for code, phase in re.findall(r"\b((?:RL|CL|SU)-\d+(?:-L\d+)?)\s*\((RD|RM|RU)\)", text, flags=re.IGNORECASE)
    ]


def build_phase_refs_from_codes(codes, phase):
    return [f"{code} ({phase})" for code in (codes or [])]


def normalize_worker_role(value):
    text = normalize_text(value)
    if not text:
        return None
    return WORKER_ROLE_ALIASES.get(text.lower())


def parse_role_counts_from_text(value, phase_code):
    text = normalize_text(value)
    if not text or "|" not in text:
        return {}

    phase_marker = f"{phase_code}:"
    segment = None
    for chunk in text.split("|"):
        candidate = normalize_text(chunk)
        if candidate and candidate.upper().startswith(phase_marker):
            segment = candidate[len(phase_marker):].strip()
            break

    if not segment:
        return {}

    role_counts = {}
    for piece in re.split(r"[;,]+", segment):
        item = normalize_text(piece)
        if not item:
            continue
        match = re.match(r"(.+?)\s+(\d+)$", item)
        if not match:
            continue
        role_id = normalize_worker_role(match.group(1))
        if not role_id:
            continue
        role_counts[role_id] = max(0, int(match.group(2)))
    return role_counts


def resolve_sheet(workbook, candidate_names):
    sheet = next((workbook[name] for name in candidate_names if name in workbook.sheetnames), None)
    if sheet is None:
        raise KeyError(f"None of the expected sheets were found: {', '.join(candidate_names)}")
    return sheet


def parse_phase_identifier(value):
    text = normalize_text(value)
    if not text:
        return None, None
    match = PHASE_ROW_PATTERN.match(text)
    if not match:
        return None, None
    return match.group(1).upper(), match.group(2).upper()


def parse_role_counts_from_columns(values, start_index):
    role_counts = {}
    for offset, role_id in enumerate(WORKBOOK_ROLE_COLUMNS):
        count = normalize_int(values[start_index + offset] if start_index + offset < len(values) else None) or 0
        if count > 0:
            role_counts[role_id] = count
    return role_counts


def group_expanded_phase_rows(sheet):
    grouped = {}

    for values in sheet.iter_rows(min_row=4, values_only=True):
        code, phase = parse_phase_identifier(values[0] if values else None)
        if not code:
            continue

        entry = grouped.setdefault(
            code,
            {
                "code": code,
                "load_type": code.split("-", 1)[0],
                "description": None,
                "category": None,
                "weight_tons": None,
                "weight_text": None,
                "dimensions": None,
                "dimensions_text": None,
                "priority": 0,
                "truck_type": None,
                "truck_types": [],
                "is_critical": False,
                "rig_down_dependency_codes": [],
                "rig_down_dependency_phase_codes": [],
                "rig_move_dependency_codes": [],
                "rig_move_dependency_phase_codes": [],
                "rig_up_dependency_codes": [],
                "rig_up_dependency_phase_codes": [],
                "avg_rig_down_minutes": None,
                "avg_rig_up_minutes": None,
                "optimal_rig_down_minutes": None,
                "optimal_rig_up_minutes": None,
                "minimum_crew_down_roles": {},
                "minimum_crew_up_roles": {},
                "optimal_crew_down_roles": {},
                "optimal_crew_up_roles": {},
            },
        )

        description = normalize_text(values[1] if len(values) > 1 else None)
        category = normalize_text(values[2] if len(values) > 2 else None)
        weight_text = normalize_text(values[3] if len(values) > 3 else None)
        dimensions_text = normalize_text(values[4] if len(values) > 4 else None)
        priority = normalize_int(values[5] if len(values) > 5 else None)
        truck_type = normalize_text(values[6] if len(values) > 6 else None)
        is_critical = parse_yes_no(values[7] if len(values) > 7 else None)
        duration_minutes = parse_duration_minutes(values[8] if len(values) > 8 else None)
        predecessor_codes = parse_dependency_codes(values[9] if len(values) > 9 else None)
        predecessor_phase_codes = parse_phase_dependency_refs(values[9] if len(values) > 9 else None)
        minimum_role_counts = parse_role_counts_from_columns(values, 10)
        optimal_role_counts = parse_role_counts_from_columns(values, 28)

        if description:
            entry["description"] = description
        if category:
            entry["category"] = category
        if weight_text:
            entry["weight_text"] = weight_text
            entry["weight_tons"] = parse_weight_tons(weight_text)
        if dimensions_text:
            entry["dimensions_text"] = dimensions_text
            entry["dimensions"] = parse_dimension_triplet(dimensions_text)
        if priority is not None:
            entry["priority"] = priority
        if truck_type:
            entry["truck_type"] = truck_type
            entry["truck_types"] = parse_truck_types(truck_type)
        entry["is_critical"] = entry["is_critical"] or is_critical

        if phase == "RD":
            entry["rig_down_dependency_codes"] = predecessor_codes
            entry["rig_down_dependency_phase_codes"] = predecessor_phase_codes
            entry["avg_rig_down_minutes"] = duration_minutes
            entry["optimal_rig_down_minutes"] = duration_minutes
            entry["minimum_crew_down_roles"] = minimum_role_counts
            entry["optimal_crew_down_roles"] = optimal_role_counts
        elif phase == "RM":
            entry["rig_move_dependency_codes"] = predecessor_codes
            entry["rig_move_dependency_phase_codes"] = predecessor_phase_codes
        elif phase == "RU":
            entry["rig_up_dependency_codes"] = predecessor_codes
            entry["rig_up_dependency_phase_codes"] = predecessor_phase_codes
            entry["avg_rig_up_minutes"] = duration_minutes
            entry["optimal_rig_up_minutes"] = duration_minutes
            entry["minimum_crew_up_roles"] = minimum_role_counts
            entry["optimal_crew_up_roles"] = optimal_role_counts

    return [grouped[key] for key in sorted(grouped.keys())]


def parse_compact_rig_load_rows(sheet):
    rows = []

    for values in sheet.iter_rows(min_row=2, values_only=True):
        code = normalize_text(values[0] if values else None)
        if not code:
            continue

        rig_down_dependency_codes = parse_dependency_codes(values[8] if len(values) > 8 else None)
        rig_up_dependency_codes = parse_dependency_codes(values[9] if len(values) > 9 else None)
        minimum_crew_text = values[14] if len(values) > 14 else None
        optimal_crew_text = values[15] if len(values) > 15 else None

        rows.append(
            {
                "code": code.upper(),
                "load_type": normalize_text(values[1] if len(values) > 1 else None) or code.split("-", 1)[0],
                "description": normalize_text(values[2] if len(values) > 2 else None),
                "category": normalize_text(values[3] if len(values) > 3 else None),
                "load_count": max(1, normalize_int(values[4] if len(values) > 4 else None) or 1),
                "weight_tons": parse_weight_tons(values[5] if len(values) > 5 else None),
                "weight_text": normalize_text(values[5] if len(values) > 5 else None),
                "dimensions": parse_dimension_triplet(values[6] if len(values) > 6 else None),
                "dimensions_text": normalize_text(values[6] if len(values) > 6 else None),
                "priority": normalize_int(values[7] if len(values) > 7 else None) or 0,
                "rig_down_dependency_codes": rig_down_dependency_codes,
                "rig_down_dependency_phase_codes": build_phase_refs_from_codes(rig_down_dependency_codes, "RD"),
                "rig_move_dependency_codes": [],
                "rig_move_dependency_phase_codes": [],
                "rig_up_dependency_codes": rig_up_dependency_codes,
                "rig_up_dependency_phase_codes": build_phase_refs_from_codes(rig_up_dependency_codes, "RU"),
                "avg_rig_down_minutes": parse_duration_minutes(values[10] if len(values) > 10 else None),
                "avg_rig_up_minutes": parse_duration_minutes(values[11] if len(values) > 11 else None),
                "is_critical": parse_yes_no(values[12] if len(values) > 12 else None),
                "truck_type": normalize_text(values[13] if len(values) > 13 else None),
                "truck_types": parse_truck_types(values[13] if len(values) > 13 else None),
                "minimum_crew_down_roles": parse_role_counts_from_text(minimum_crew_text, "RD"),
                "minimum_crew_up_roles": parse_role_counts_from_text(minimum_crew_text, "RU"),
                "optimal_crew_down_roles": parse_role_counts_from_text(optimal_crew_text, "RD"),
                "optimal_crew_up_roles": parse_role_counts_from_text(optimal_crew_text, "RU"),
                "optimal_rig_down_minutes": parse_duration_minutes(values[16] if len(values) > 16 else None) or parse_duration_minutes(values[10] if len(values) > 10 else None),
                "optimal_rig_up_minutes": parse_duration_minutes(values[17] if len(values) > 17 else None) or parse_duration_minutes(values[11] if len(values) > 11 else None),
            }
        )

    return rows


def parse_compact_startup_load_rows(sheet):
    rows = []

    for values in sheet.iter_rows(min_row=2, values_only=True):
        code = normalize_text(values[0] if values else None)
        if not code:
            continue

        dependency_codes = parse_dependency_codes(values[7] if len(values) > 7 else None)
        dependency_phase_codes = build_phase_refs_from_codes(dependency_codes, "RU")
        avg_rig_up_minutes = parse_duration_minutes(values[8] if len(values) > 8 else None)

        rows.append(
            {
                "code": code.upper(),
                "load_type": normalize_text(values[1] if len(values) > 1 else None) or "Startup",
                "description": normalize_text(values[2] if len(values) > 2 else None),
                "count": max(1, normalize_int(values[3] if len(values) > 3 else None) or 1),
                "weight_tons": parse_weight_tons(values[4] if len(values) > 4 else None),
                "weight_text": normalize_text(values[4] if len(values) > 4 else None),
                "dimensions": parse_dimension_triplet(values[5] if len(values) > 5 else None),
                "dimensions_text": normalize_text(values[5] if len(values) > 5 else None),
                "priority": normalize_int(values[6] if len(values) > 6 else None) or 0,
                "dependencyLabel": ", ".join(dependency_codes) if dependency_codes else "Standalone startup load",
                "rig_move_dependency_codes": dependency_codes,
                "rig_move_dependency_phase_codes": dependency_phase_codes,
                "rig_up_dependency_codes": dependency_codes,
                "rig_up_dependency_phase_codes": dependency_phase_codes,
                "avg_rig_up_minutes": avg_rig_up_minutes,
                "truck_type": normalize_text(values[9] if len(values) > 9 else None),
                "truck_types": parse_truck_types(values[9] if len(values) > 9 else None),
                "isReusable": code.upper() in STARTUP_REUSABLE_IDS,
                "minimum_crew_up_count": 0,
                "optimal_crew_up_count": 0,
                "minimum_crew_up_roles": {},
                "optimal_crew_up_roles": {},
            }
        )

    return rows


def build_rig_load_payload(row):
    return {
        "id": row["code"],
        "code": row["code"],
        "load_type": row.get("load_type"),
        "description": row.get("description"),
        "category": row.get("category"),
        "load_count": 1,
        "weight_tons": row.get("weight_tons"),
        "weight_text": row.get("weight_text"),
        "dimensions": row.get("dimensions"),
        "dimensions_text": row.get("dimensions_text"),
        "priority": row.get("priority") or 0,
        "rig_down_dependency_codes": row.get("rig_down_dependency_codes") or [],
        "rig_down_dependency_phase_codes": row.get("rig_down_dependency_phase_codes") or [],
        "rig_move_dependency_codes": row.get("rig_move_dependency_codes") or [],
        "rig_move_dependency_phase_codes": row.get("rig_move_dependency_phase_codes") or [],
        "rig_up_dependency_codes": row.get("rig_up_dependency_codes") or [],
        "rig_up_dependency_phase_codes": row.get("rig_up_dependency_phase_codes") or [],
        "avg_rig_down_minutes": row.get("avg_rig_down_minutes"),
        "avg_rig_up_minutes": row.get("avg_rig_up_minutes"),
        "is_critical": bool(row.get("is_critical")),
        "truck_type": row.get("truck_type"),
        "truck_types": row.get("truck_types") or [],
        "minimum_crew_down_count": sum((row.get("minimum_crew_down_roles") or {}).values()),
        "minimum_crew_up_count": sum((row.get("minimum_crew_up_roles") or {}).values()),
        "optimal_crew_down_count": sum((row.get("optimal_crew_down_roles") or {}).values()),
        "optimal_crew_up_count": sum((row.get("optimal_crew_up_roles") or {}).values()),
        "minimum_crew_down_roles": row.get("minimum_crew_down_roles") or {},
        "minimum_crew_up_roles": row.get("minimum_crew_up_roles") or {},
        "optimal_crew_down_roles": row.get("optimal_crew_down_roles") or {},
        "optimal_crew_up_roles": row.get("optimal_crew_up_roles") or {},
        "optimal_rig_down_minutes": row.get("optimal_rig_down_minutes"),
        "optimal_rig_up_minutes": row.get("optimal_rig_up_minutes"),
    }


def build_startup_load_payload(row):
    return {
        "id": row["code"],
        "code": row["code"],
        "load_type": row.get("load_type"),
        "description": row.get("description"),
        "count": 1,
        "weight_tons": row.get("weight_tons"),
        "weight_text": row.get("weight_text"),
        "dimensions": row.get("dimensions"),
        "dimensions_text": row.get("dimensions_text"),
        "priority": row.get("priority") or 0,
        "dependencyLabel": ", ".join(row.get("rig_move_dependency_codes") or row.get("rig_up_dependency_codes") or []) or "Standalone startup load",
        "rig_move_dependency_codes": row.get("rig_move_dependency_codes") or [],
        "rig_move_dependency_phase_codes": row.get("rig_move_dependency_phase_codes") or [],
        "rig_up_dependency_codes": row.get("rig_up_dependency_codes") or [],
        "rig_up_dependency_phase_codes": row.get("rig_up_dependency_phase_codes") or [],
        "avg_rig_up_minutes": row.get("avg_rig_up_minutes"),
        "truck_type": row.get("truck_type"),
        "truckTypes": row.get("truck_types") or [],
        "isReusable": row["code"] in STARTUP_REUSABLE_IDS,
        "minimum_crew_up_count": sum((row.get("minimum_crew_up_roles") or {}).values()),
        "optimal_crew_up_count": sum((row.get("optimal_crew_up_roles") or {}).values()),
        "minimum_crew_up_roles": row.get("minimum_crew_up_roles") or {},
        "optimal_crew_up_roles": row.get("optimal_crew_up_roles") or {},
    }


def build_truck_spec_payload(values):
    truck_type = normalize_truck_type(values[0] if len(values) > 0 else None)
    if not truck_type:
        return None

    alpha_values = [float(match) for match in re.findall(r"\d+(?:\.\d+)?", normalize_text(values[4] if len(values) > 4 else None) or "")]
    alpha = sum(alpha_values) / len(alpha_values) if alpha_values else 0.3

    return {
        "type": truck_type,
        "max_weight_tons": normalize_float(values[1] if len(values) > 1 else None) or 0,
        "dimensions": parse_dimension_triplet(values[2] if len(values) > 2 else None),
        "average_speed_kmh": normalize_float(values[3] if len(values) > 3 else None) or 0,
        "alpha": alpha,
    }


def fits_load_to_truck_type(load, truck_spec):
    if not truck_spec:
        return False

    dimensions = load.get("dimensions") or {}
    truck_dimensions = truck_spec.get("dimensions") or {}

    weight_ok = (
        not load.get("weight_tons")
        or not truck_spec.get("max_weight_tons")
        or float(load["weight_tons"]) <= float(truck_spec["max_weight_tons"])
    )
    length_ok = (
        not dimensions.get("length")
        or not truck_dimensions.get("length")
        or float(dimensions["length"]) <= float(truck_dimensions["length"])
    )
    width_ok = (
        not dimensions.get("width")
        or not truck_dimensions.get("width")
        or float(dimensions["width"]) <= float(truck_dimensions["width"])
    )
    height_ok = (
        not dimensions.get("height")
        or not truck_dimensions.get("height")
        or float(dimensions["height"]) <= float(truck_dimensions["height"])
    )

    return weight_ok and length_ok and width_ok and height_ok


def can_move_as_oversize_permit(load, truck_spec):
    if not truck_spec:
        return False

    truck_type = normalize_text(truck_spec.get("type")) or ""
    if truck_type not in {"Low-bed", "Heavy Hauler"}:
        return False

    load_weight = float(load.get("weight_tons") or 0)
    max_weight = float(truck_spec.get("max_weight_tons") or 0)
    if load_weight and max_weight and load_weight > max_weight:
        return False

    return True


def augment_feasible_truck_types(loads, truck_specs):
    ordered_specs = sorted(
        (truck_specs or []),
        key=lambda spec: (
            float(spec.get("max_weight_tons") or 0),
            float((spec.get("dimensions") or {}).get("length") or 0),
            float((spec.get("dimensions") or {}).get("width") or 0),
            float((spec.get("dimensions") or {}).get("height") or 0),
        ),
    )

    for load in loads or []:
        declared_types = list(load.get("truck_types") or [])
        feasible_declared = [
            truck_type
            for truck_type in declared_types
            if (
                fits_load_to_truck_type(load, next((spec for spec in ordered_specs if spec.get("type") == truck_type), None))
                or can_move_as_oversize_permit(load, next((spec for spec in ordered_specs if spec.get("type") == truck_type), None))
            )
        ]

        if feasible_declared:
          load["truck_types"] = feasible_declared
          load["truck_type"] = " / ".join(feasible_declared)
          continue

        feasible_fallbacks = [
            spec.get("type")
            for spec in ordered_specs
            if fits_load_to_truck_type(load, spec) or can_move_as_oversize_permit(load, spec)
        ]
        if feasible_fallbacks:
            load["truck_types"] = feasible_fallbacks
            load["truck_type"] = " / ".join(feasible_fallbacks)

    return loads


@lru_cache(maxsize=1)
def load_planning_dataset():
    workbook = load_workbook(DATASET_PATH, data_only=True)
    rig_sheet = resolve_sheet(workbook, SHEET_NAME_CANDIDATES["rig"])
    startup_sheet = resolve_sheet(workbook, SHEET_NAME_CANDIDATES["startup"])

    grouped_rig_rows = group_expanded_phase_rows(rig_sheet)
    grouped_startup_rows = group_expanded_phase_rows(startup_sheet)

    if grouped_rig_rows:
        rig_loads = [build_rig_load_payload(row) for row in grouped_rig_rows]
    else:
        rig_loads = parse_compact_rig_load_rows(rig_sheet)

    if grouped_startup_rows:
        startup_loads = [build_startup_load_payload(row) for row in grouped_startup_rows]
    else:
        startup_loads = parse_compact_startup_load_rows(startup_sheet)

    truck_specs = [
        payload
        for payload in (
            build_truck_spec_payload(values)
            for values in resolve_sheet(workbook, SHEET_NAME_CANDIDATES["truck"]).iter_rows(min_row=2, values_only=True)
        )
        if payload
    ]
    rig_loads = augment_feasible_truck_types(rig_loads, truck_specs)
    startup_loads = augment_feasible_truck_types(startup_loads, truck_specs)

    return {
        "rig_loads": rig_loads,
        "startup_loads": startup_loads,
        "truck_specs": truck_specs,
    }


def get_worker_roles():
    return WORKER_ROLE_DEFINITIONS
